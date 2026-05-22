import base64
import io
import json
from pathlib import Path
from PIL import Image
from langchain_ollama import ChatOllama
from langchain_core.messages import HumanMessage
from pydantic import BaseModel
from typing import Literal
from langchain_core.messages import SystemMessage
import openpyxl
from openpyxl import load_workbook
import re
from config import CARTELLA_IMMAGINI, FILE_EXCEL
#nome del modello preso da ollama
MODELLO           = "ministral-3:3b"

class ValutazioneDisegno(BaseModel):
    tipologia:    Literal["+", "-", "+-"]                             
    dimensione:   Literal["L", "S", "N", "SL"]                        
    colore:       Literal["R-D", "R-L", "NR-D", "NR-L"]               
    posizione:    Literal["C", "P-L", "P-LT", "P-LB", "P-R", "P-RT", "P-RB"] 
    orientamento: Literal["S", "U", "D"]                               
    accuratezza:  Literal["A-D", "NA-D", "A-ND", "NA-ND"]            
    alim:         bool                                                  
    io:           bool                                                 
    sport:        bool                                                 
    assoc:        Literal["++", "--", "=="]                      

SYSTEM_PROMPT = """Sei un analista esperto di disegni infantili in ambito sanitario.
Il tuo compito è analizzare disegni fatti da bambini che hanno ricevuto educazione su:
- alimentazione sana (cibi sani vs batteri/cibi non sani)
- igiene orale (denti, spazzolino, carie)
- attività fisica vs sedentarietà

REGOLA FONDAMENTALE: per ogni campo devi scegliere SOLO uno dei valori esatti indicati.
Non inventare valori nuovi. Se non sei sicuro, scegli il valore più vicino tra quelli disponibili.
Non usare null, N/A o altri placeholder — scegli sempre un valore valido.

Devi rispondere ESCLUSIVAMENTE con un oggetto JSON valido, senza testo aggiuntivo, 
senza markdown, senza backtick. Solo JSON puro.
Esempio valido:

{
  "tipologia": "+",
  "dimensione": "N",
  "colore": "R-L",
  "posizione": "C",
  "orientamento": "S",
  "accuratezza": "A-ND",
  "alim": true,
  "io": true,
  "sport": false,
  "assoc": "=="
}

"""

USER_PROMPT = """
Restituisci SOLO JSON valido.
Schema:

{
  "tipologia": "+ | - | +-",
  "dimensione": "L | S | N | SL",
  "colore": "R-D | R-L | NR-D | NR-L",
  "posizione": "C | P-L | P-LT | P-LB | P-R | P-RT | P-RB",
  "orientamento": "S | U | D",
  "accuratezza": "A-D | NA-D | A-ND | NA-ND",
  "alim": true/false,
  "io": true/false,
  "sport": true/false,
  "assoc": "++ | -- | =="
}


Legenda valori:
- tipologia: "+" elementi positivi (cibo sano, dente sano), "-" elementi negativi (batteri, cibo cattivo), "+-" misti
- dimensione: L grande, S piccola, N normale, SL sproporzione tra elementi
- colore: R-D uso realistico del colore e colori scuri (es grigio, nero), R-L realistico chiaro, NR-D non realistico scuro, NR-L non realistico chiaro
- posizione: posizione prevalente degli elementi nel foglio (C centro, P–L = periferic left, a sinistra del foglio, P–LT = periferic left top, a sinistra in alto del foglio, P–LB = periferic left bottom, a sinistra in basso del foglio, P–R = periferic right, a destra del foglio, P–RT = periferic right top, a destra in alto del foglio, P–RB = periferic right bottom, a destra in basso del foglio)
- orientamento: S dritto, U sottosopra, D distorto
- accuratezza: A-D accurato e dettagliato, NA-D non accurato ma dettagliato, A-ND accurato non dettagliato, NA-ND non accurato non dettagliato
- alim: true se ci sono alimenti nel disegno
- io: true se ci sono denti o strumenti per igiene orale
- sport: true se c'è attività fisica o televisione/sedentarietà
- assoc: Associazione fra gli elementi presenti nel disegno (es. cibi sani e dente felice; oppure dente cariato e batteri, ecc.); se sì mettere associazione corretta o non forte e non evidente ++, --, == se non c'è nessuna associazione"""

#funzione di conversione da .tif a .png in base64 per il modello
def load_image_as_base64(path: Path) -> str:
    try:
        with Image.open(path) as img:
            img.load()  # forza il caricamento completo prima della conversione
            img = img.convert("RGB")
            buffer = io.BytesIO()
            img.save(buffer, format="PNG")
            return base64.b64encode(buffer.getvalue()).decode("utf-8")
    except Exception as e:
        raise ValueError(f"Impossibile leggere il file TIF '{path.name}': {e}")

#funzione per creare file di excel o caricarlo se esiste già
def init_excel(path: Path) -> openpyxl.Workbook:
    if path.exists():
        return load_workbook(path)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valutazioni AI"
    intestazioni = [
        "disegno_id", 
        "tipologia", "dimensione", "colore", "posizione",
        "orientamento", "accuratezza", "alim", "io", "sport",
        "assoc"
    ]
    ws.append(intestazioni)
    wb.save(path)
    return wb

#funzione di lettura del file excel per non rianalizzare disegni già analizzati nel caso di interruzioni
def get_disegni_gia_analizzati(wb: openpyxl.Workbook) -> set:
    ws = wb.active
    analizzati = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            analizzati.add(int(row[0]))
    return analizzati

#funzione per scrivere una riga nel file excel
def scrivi_riga(wb: openpyxl.Workbook, path: Path, disegno_id: int, 
                risultato: ValutazioneDisegno = None, errore: str = None):
    ws = wb.active
    if risultato:
        ws.append([
            disegno_id,
            risultato.tipologia, risultato.dimensione, risultato.colore,
            risultato.posizione, risultato.orientamento, risultato.accuratezza,
            risultato.alim, risultato.io, risultato.sport,
            risultato.assoc,
        ])
    else:
        ws.append([disegno_id,
                   None, None, None, None, None, None,
                   None, None, None, None, None, None])
    wb.save(path)

#funzione per estrarre il json della risposta del modello
def estrai_json(testo: str) -> str:
    testo = testo.strip()

    match = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", testo, re.DOTALL)
    if match:
        testo = match.group(1)
    else:
        match = re.search(r"\{.*\}", testo, re.DOTALL)
        if match:
            testo = match.group(0)
        else:
            raise ValueError(f"Nessun JSON trovato nella risposta: {repr(testo[:200])}")

    testo = re.sub(r",\s*([}\]])", r"\1", testo)
    
    return testo

def main():
    llm = ChatOllama(model=MODELLO, temperature=0, keep_alive=0)
    wb  = init_excel(FILE_EXCEL)
    
    #Recupera id già analizzati (per riprendere dopo interruzioni)
    gia_analizzati = get_disegni_gia_analizzati(wb)
    if gia_analizzati:
        print(f"Trovati {len(gia_analizzati)} disegni già analizzati, li salto.")

    #Raccoglie tutti i .tif ordinati per numero
    immagini = sorted(
        CARTELLA_IMMAGINI.glob("*.tif"),
        key=lambda p: int(p.stem) if p.stem.isdigit() else 0
    )
    
    totale  = len(immagini)
    errori  = 0

    for i, img_path in enumerate(immagini, 1):
        disegno_id = int(img_path.stem)

        # Salta se già analizzato
        if disegno_id in gia_analizzati:
            continue

        print(f"[{i}/{totale}] Analisi disegno {disegno_id}...", end=" ")

        try:
            image_data = load_image_as_base64(img_path)

            msg = HumanMessage(content=[
                {"type": "text",      "text": USER_PROMPT},
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_data}"}}
            ])

            response = llm.invoke([SystemMessage(content=SYSTEM_PROMPT),msg])
            
            testo     = response.content.strip()
            print(f"DEBUG risposta raw: {repr(testo[:300])}")
            # Pulisce eventuali backtick residui
            testo = estrai_json(testo)

            dati      = json.loads(testo)
            risultato = ValutazioneDisegno(**dati)

            scrivi_riga(wb, FILE_EXCEL, disegno_id, risultato=risultato)
            print(f"OK")

        except json.JSONDecodeError as e:
            errori += 1
            msg_err = f"JSON non valido: {e}"
            scrivi_riga(wb, FILE_EXCEL, disegno_id, errore=msg_err)
            print(f"ERRORE JSON: {msg_err}")

        except Exception as e:
            errori += 1
            msg_err = str(e)
            scrivi_riga(wb, FILE_EXCEL, disegno_id, errore=msg_err)
            print(f"ERRORE: {msg_err}")

    print(f"\nCompletato. {totale - errori}/{totale} disegni analizzati con successo.")

main()