import pandas as pd
from sklearn.metrics import confusion_matrix, classification_report
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from config import TABELLA_PSICOLOGA, TABELLA_CONFRONTO, MATRICE_CONFUSIONE


def write_str(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=str(value) if value is not None else "")
    cell.data_type = 's'

def write_confusion_sheet(wb: Workbook, col_name: str,
                          y_true: pd.Series, y_pred: pd.Series):
    
    labels = sorted(set(y_true) | set(y_pred))
    cm = confusion_matrix(y_true, y_pred, labels=labels)
    report = classification_report(
        y_true, y_pred, labels=labels,
        output_dict=True, zero_division=0
    )

    ws = wb.create_sheet(title=col_name[:31])

    ws.cell(row=1, column=1,
            value=f"Matrice di Confusione – {col_name}  [riga=true / colonna=predicted]")

    ws.cell(row=3, column=1, value="True \\ Predicted")
    for j, lbl in enumerate(labels):
        write_str(ws, 3, j + 2, lbl)
    ws.cell(row=3, column=len(labels) + 2, value="Totale")

    for i, lbl_row in enumerate(labels):
        r = i + 4
        write_str(ws, i + 4, 1, lbl_row)
        for j in range(len(labels)):
            ws.cell(row=r, column=j + 2, value=int(cm[i, j]))
        ws.cell(row=r, column=len(labels) + 2, value=int(cm[i].sum()))

    tot_r = len(labels) + 4
    ws.cell(row=tot_r, column=1, value="Totale")
    for j in range(len(labels)):
        ws.cell(row=tot_r, column=j + 2, value=int(cm[:, j].sum()))
    ws.cell(row=tot_r, column=len(labels) + 2, value=int(cm.sum()))

    ms = tot_r + 2
    ws.append([]) 
    headers = ["Classe", "Precision", "Recall"]
    for k, h in enumerate(headers):
        ws.cell(row=ms, column=k + 1, value=h)

    for i, lbl in enumerate(labels):
        r = ms + 1 + i
        d = report.get(str(lbl), {})
        write_str(ws, ms + 1 + i, 1, lbl)
        ws.cell(row=r, column=2, value=round(d.get("precision", 0), 4))
        ws.cell(row=r, column=3, value=round(d.get("recall",    0), 4))

    acc_r = ms + 1 + len(labels)
    ws.cell(row=acc_r, column=1, value="Accuracy")
    ws.cell(row=acc_r, column=2, value=round(report.get("accuracy", 0), 4))

    for offset, key in [(1, "macro avg"), (2, "weighted avg")]:
        r = acc_r + offset
        d = report.get(key, {})
        ws.cell(row=r, column=1, value=key.title())
        ws.cell(row=r, column=2, value=round(d.get("precision", 0), 4))
        ws.cell(row=r, column=3, value=round(d.get("recall",    0), 4))

    ws.column_dimensions["A"].width = 22
    for j in range(1, len(labels) + 3):
        ws.column_dimensions[get_column_letter(j + 1)].width = 14


def write_missing_sheet(wb: Workbook, missing1: pd.DataFrame, missing2: pd.DataFrame):
    ws = wb.create_sheet(title="Dati Mancanti")

    def block(start_row, title, df):
        ws.cell(row=start_row, column=1, value=title)
        if df.empty:
            ws.cell(row=start_row + 1, column=1, value="Nessuna riga con dati mancanti")
            return start_row + 3
        for j, col in enumerate(df.columns):
            ws.cell(row=start_row + 1, column=j + 1, value=col)
        for i, (_, row) in enumerate(df.iterrows()):
            for j, col in enumerate(df.columns):
                ws.cell(row=start_row + 2 + i, column=j + 1, value=row[col])
        return start_row + 2 + len(df) + 2

    r = block(1, "Tabella 1 – ground truth", missing1)
    block(r,     "Tabella 2 – predicted",    missing2)


def check_missing(df: pd.DataFrame, id_col: str) -> pd.DataFrame:
    cols = [c for c in df.columns if c != id_col]
    mask = df[cols].isnull().any(axis=1) | (df[cols].astype(str) == "").any(axis=1)
    return df[mask][[id_col] + cols]


def main():
    df1 = pd.read_excel(TABELLA_PSICOLOGA, dtype=str, engine="calamine")
    df2 = pd.read_excel(TABELLA_CONFRONTO, dtype=str, engine="calamine")
    
    df1.columns = df2.columns = [str(c).strip() for c in df1.columns]

    id_col      = df1.columns[0]
    target_cols = df1.columns[1:].tolist()

    df1[id_col] = df1[id_col].str.strip()
    df2[id_col] = df2[id_col].str.strip()

    merged = pd.merge(df1, df2, on=id_col, how="inner", suffixes=("_true", "_pred"))

    wb = Workbook()
    wb.remove(wb.active)

    for col in target_cols:
        subset = merged[[f"{col}_true", f"{col}_pred"]].dropna()
        y_true = subset[f"{col}_true"].str.strip()
        y_pred = subset[f"{col}_pred"].str.strip()
        write_confusion_sheet(wb, col, y_true, y_pred)

    write_missing_sheet(wb, check_missing(df1, id_col), check_missing(df2, id_col))

    wb.save(MATRICE_CONFUSIONE)
    print(f"Salvato: {MATRICE_CONFUSIONE}")



main()
